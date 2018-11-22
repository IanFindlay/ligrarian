#!/usr/bin/env python3

"""Do a test run of the Ligrarian that resets everything afterwards."""

from datetime import datetime as dt

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

import ligrarian


email = ligrarian.get_setting('User', 'Email')
password = ligrarian.get_setting('User', 'Password')
today = dt.strftime(dt.now(), '%d/%m/%y')
book_info = {
    'title': 'Cannery Row', 'author': 'John Steinbeck', 'date': today,
    'format': 'kindle', 'rating': '4',  'review': 'Test Review',
}

driver = webdriver.Firefox()
driver.implicitly_wait(10)

ligrarian.goodreads_login(driver, email, password)
url = ligrarian.goodreads_find(driver, book_info)

# Check book is correct format
info_rows = driver.find_elements_by_class_name('row')
row_text = [row.text for row in info_rows]
assert 'Kindle' in ''.join(row_text), "Book is in incorrect format."

shelves_list = ligrarian.goodreads_update(driver, book_info)

undo_elem = driver.find_element_by_class_name('wtrStatusRead.wtrUnshelve')
assert undo_elem, "Book wasn't marked as read."
print("Book successfully marked as read.")
# Reset Goodreads account
undo_elem.click()
alert_obj = driver.switch_to.alert
alert_obj.accept()

# Check book no longer marked as read before closing window
unread_check = driver.find_element_by_class_name('wtrToRead')
print('Goodreads account reset correctly.')
driver.close()

# Spreadsheet entry testing
path = ligrarian.get_setting('Settings', 'Path')
wb = openpyxl.load_workbook(path)
print('Testing spreadsheet updating')
info = ligrarian.parse_page(url, shelves_list)

year_sheet = '20' + book_info['date'][-2:]

# Get first blank row before attempting to write
pre_year = ligrarian.first_blank(wb[year_sheet])
pre_overall = ligrarian.first_blank(wb['Overall'])
wb.close()

# Try to write info to both sheets
ligrarian.input_info(year_sheet, info, book_info['date'])

# Find blank rows now that the data should have been entered
wb = openpyxl.load_workbook(path)
post_year = ligrarian.first_blank(wb[year_sheet])
post_overall = ligrarian.first_blank(wb['Overall'])

# Confirm data was entered
assert pre_year < post_year, "Data not written to year sheet."
assert pre_overall < post_overall, "Data not written to overall sheet."
print("Data was written to sheet successfully.")

# Delete newly entered data
for sheet in [year_sheet, 'Overall']:
    sheet = wb[sheet]

    input_row = ligrarian.first_blank(sheet) - 1

    sheet.cell(row=input_row, column=1).value = ''
    sheet.cell(row=input_row, column=2).value = ''
    sheet.cell(row=input_row, column=3).value = ''
    sheet.cell(row=input_row, column=4).value = ''
    sheet.cell(row=input_row, column=5).value = ''
    sheet.cell(row=input_row, column=6).value = ''

wb.save(path)

print("Spreadsheet reset.")
print("Test run complete.")
