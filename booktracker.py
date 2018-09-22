#!/usr/bin/env python3

"""Automatically update Goodreads and local Spreadsheet with book read info."""

import configparser
import datetime
from datetime import timedelta
import time
import tkinter as tk
import tkinter.messagebox
import sys

import bs4
import openpyxl
import requests
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select


class GuiInput:
    """Booktracker GUI structure and related methods."""

    def __init__(self, master):
        self.master = master
        master.title("Booktracker")
        width = 680
        height = 470
        geometry_string = "{}x{}".format(width, height)
        root.geometry(geometry_string)

        self.title_label = tk.Label(root, text="Book Title".ljust(20))
        self.title_label.grid(row=1, column=1, sticky='W')

        self.author_label = tk.Label(root, text="Author Name")
        self.author_label.grid(row=2, column=1, sticky='W')

        self.format_label = tk.Label(root, text="Format")
        self.format_label.grid(row=3, column=1, sticky='W')

        self.rating_label = tk.Label(root, text="Rating")
        self.rating_label.grid(row=4, column=1, sticky='W')

        self.date_label = tk.Label(root, text="Date Read")
        self.date_label.grid(row=5, column=1, sticky='W')

        self.review_label = tk.Label(root, text="Review (optional)")
        self.review_label.grid(row=6, column=1, sticky='W')

        self.title = tk.Entry(root)
        self.title.grid(row=1, column=2, sticky='W', pady=5)

        self.author = tk.Entry(root)
        self.author.grid(row=2, column=2, sticky='w', pady=5)

        formats = ("Ebook", "Hardback", "Kindle", "Paperback",)
        self.book_format = tk.StringVar()
        format_drop = tk.OptionMenu(root, self.book_format, *formats)
        format_drop.grid(row=3, column=2, sticky='w')

        stars = ("1", "2", "3", "4", "5")
        self.star = tk.StringVar()
        rating = tk.OptionMenu(root, self.star, *stars)
        rating.grid(row=4, column=2, sticky='w')

        dates = ("Today", "Yesterday", "Custom")
        self.date = tk.StringVar()
        date_read = tk.OptionMenu(root, self.date, *dates)
        date_read.grid(row=5, column=2, sticky='w')

        self.review = tk.Text(root, height=15, width=75, wrap=tk.WORD)
        self.review.grid(row=6, column=2, sticky='e', pady=5)

        submit_button = tk.Button(root, text="Mark as Read",
                                  command=self.parse_input)
        submit_button.grid(row=12, column=2, sticky='e', pady=20)


    def parse_input(self):
        """Create and verify the book details inputted to the GUI."""
        book_info['title'] = self.title.get()
        book_info['author'] = self.author.get()
        book_info['format'] = self.book_format.get()
        book_info['rating'] = self.star.get()
        book_info['date'] = self.date.get()
        book_info['review'] = self.review.get('1.0', 'end-1c')

        try:
            assert book_info['title'] != ''
            assert book_info['author'] != ''
            assert book_info['format'] != ''
            assert book_info['rating'] != ''
            assert book_info['date'] != ''
            print(book_info)
            root.destroy()

        except AssertionError:
            tk.messagebox.showwarning(message="Complete all non-optional "
                                          "fields before marking as read")


def get_date(book_info):
    """Return the date the book was read formatted (DD/MM/YY)."""
    today = datetime.datetime.now()

    if book_info['date'] == 'Today':
        date = today.strftime('%d/%m/%y')

    elif book_info['date'] == 'Yesterday':
        yesterday = today - timedelta(days=1)
        date = yesterday.strftime('%d/%m/%y')

    else:
        date = input('Enter the date the book was finished (DD/MM/YY): ')

    return date


def goodreads_find(book_info, username, password):
    """Find the correct book, in the correct format, on Goodreads."""
    driver.get('https://goodreads.com')

    # Login
    email_elem = driver.find_element_by_name('user[email]')
    email_elem.send_keys(username)
    pass_elem = driver.find_element_by_name('user[password]')
    pass_elem.send_keys(password)
    pass_elem.send_keys(Keys.ENTER)

    # Find correct book and edition
    book_title = book_info['title']
    book_author = book_info['author']
    search_elem = driver.find_element_by_class_name('searchBox__input')
    search_elem.send_keys(book_title + ' ' + book_author + '%3Dauthor')
    search_elem.send_keys(Keys.ENTER)

    driver.find_element_by_class_name('bookTitle').click()

    driver.find_element_by_class_name('otherEditionsLink').click()

    # Format
    book_format = book_info['format'].lower()
    filter_elem = driver.find_element_by_name('filter_by_format')
    filter_elem.click()
    filter_elem.send_keys(book_format)
    filter_elem.send_keys(Keys.ENTER)
    time.sleep(2)

    driver.find_element_by_class_name('bookTitle').click()

    return driver.current_url


def goodreads_update(book_info):
    """Update Goodreads by marking book as read and adding information."""
    # Mark as Read
    menu_elem = driver.find_element_by_class_name('wtrRight.wtrUp')
    menu_elem.click()
    time.sleep(1)
    search_elem = driver.find_element_by_class_name('wtrShelfSearchField')
    search_elem.click()
    search_elem.send_keys('read', Keys.ENTER)
    time.sleep(1)

    # Date Selection
    year = '20' + date_finished[6:]
    month = date_finished[3:5].lstrip('0')
    day = date_finished[:2].lstrip('0')

    year_class = 'rereadDatePicker.smallPicker.endYear'
    month_class = 'rereadDatePicker.largePicker.endMonth'
    day_class = 'rereadDatePicker.smallPicker.endDay'
    Select(driver.find_element_by_class_name(year_class)
           ).select_by_visible_text(year)

    Select(driver.find_element_by_class_name(month_class)
           ).select_by_value(month)

    Select(driver.find_element_by_class_name(day_class)
           ).select_by_visible_text(day)

    # Write review if one entered
    if book_info['review']:
        review_elem = driver.find_element_by_name('review[review]')
        review_elem.click()
        review_elem.send_keys(book_info['review'])

    # Save
    save_elem = driver.find_element_by_name('next')
    save_elem.click()

    # Shelf selection
    shelves_elems = driver.find_elements_by_class_name('actionLinkLite.'
                                                       'bookPageGenreLink')
    shelves = []
    for shelf in shelves_elems:
        if ' users' not in shelf.text and shelf.text not in shelves:
            shelves.append(shelf.text)

    if book_info['rating'] == '5':
        shelves.append('5-star-books')

    time.sleep(1)
    menu_elem = driver.find_element_by_class_name('wtrRight.wtrDown')
    menu_elem.click()
    time.sleep(1)
    shelf_elem = driver.find_element_by_class_name('wtrShelfSearchField')

    for i in range(len(shelves)):
        shelf_elem.send_keys(shelves[i], Keys.ENTER)
        shelf_elem.send_keys(Keys.SHIFT, Keys.HOME, Keys.DELETE)

    menu_elem.click()
    time.sleep(1)

    # Give star rating
    rating = book_info['rating']
    stars_elem = driver.find_elements_by_class_name('star.off')
    for stars in stars_elem:
        if stars.text.strip() == '{} of 5 stars'.format(rating):
            stars.click()
            break

    return shelves


def parse_page(url):
    """Parse and return page information needed for spreadsheet."""
    info = {}
    res = requests.get(url)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    title_elem = soup.select('#bookTitle')

    rough_title = title_elem[0].getText().strip().split('\n')
    if len(rough_title) == 1:
        title = rough_title[0].strip()
    else:
        title = rough_title[0].strip() + ' ' + rough_title[2].strip()

    info['title'] = title

    author_elem = soup.select('.authorName')
    author = author_elem[0].getText().strip()
    info['author'] = author

    pages_elem = soup.findAll('span', attrs={'itemprop': 'numberOfPages'})
    pages = int(pages_elem[0].getText().strip(' pages'))
    info['pages'] = pages

    if shelves_list[0] == 'Fiction' or shelves_list[0] == 'Nonfiction':
        category = shelves_list[0]
        genre = shelves_list[1]
    else:
        genre = shelves_list[0]
        if 'Nonfiction' in shelves_list:
            category = 'Nonfiction'
        else:
            category = 'Fiction'

    info['category'] = category
    info['genre'] = genre

    return info


def input_info(sheet_name, info):
    """Write the book information to the first blank row on the given sheet."""
    sheet_names = wb.sheetnames
    if sheet_name not in sheet_names:
        create_sheet(sheet_name, sheet_names[-1])

    sheet = wb[sheet_name]

    input_row = first_blank(sheet)

    sheet.cell(row=input_row, column=1).value = info['title']
    sheet.cell(row=input_row, column=2).value = info['author']
    sheet.cell(row=input_row, column=3).value = info['pages']
    sheet.cell(row=input_row, column=4).value = info['category']
    sheet.cell(row=input_row, column=5).value = info['genre']
    sheet.cell(row=input_row, column=6).value = date_finished


def first_blank(sheet):
    """Return the first blank row of the given sheet."""
    input_row = 1
    data = ''
    while data is not None:
        data = sheet.cell(row=input_row, column=1).value
        input_row += 1
    input_row -= 1
    return input_row


def create_sheet(sheet_name, last_sheet):
    """Create a new sheet by copying and modifying the last one."""
    sheet = wb.copy_worksheet(wb[last_sheet])
    sheet.title = sheet_name
    last_row = first_blank(sheet)
    while last_row > 1:
        for col in range(1, 7):
            sheet.cell(row=last_row, column=col).value = None
        last_row -= 1
    day_tracker = '=(TODAY()-DATE({},1,1))/7'.format(sheet_name)
    sheet.cell(row=5, column=9).value = day_tracker


def user_info():
    """Retrieve user info from config file if present or prompt for info."""
    config = configparser.ConfigParser()
    config.read('settings.ini')
    username = config.get('User', 'Username')
    password = config.get('User', 'Password')

    if username == "" or password == "":
        username = input('Enter your Username here: ')
        password = input('Enter your Password here: ')
        save = input("Save this information for future use?(y/n): ")
        if save.lower() == 'y':
            with open('settings.ini', 'w') as f:
                f.write('[User]\n')
                f.write('Username = ' + username + '\n')
                f.write('Password = ' + password + '\n')

    return (username, password)


if __name__ == '__main__':

    username, password = user_info()

    if len(sys.argv) == 6:
        book_info = {
        'title': sys.argv[1], 'author': sys.argv[2], 'format': sys.argv[3],
        'rating': sys.argv[4], 'date': sys.argv[5],
        }
    else:
        book_info = {}
        root = tk.Tk()
        gui = GuiInput(root)
        root.mainloop()

    date_finished = get_date(book_info)

    print('Opening a computer controlled browser and updating Goodreads...')
    driver = webdriver.Firefox()
    driver.implicitly_wait(10)

    url = goodreads_find(book_info, username, password)
    shelves_list = goodreads_update(book_info)
    driver.close()
    print('Goodreads account updated.')

    wb = openpyxl.load_workbook('Booktracker.xlsx')
    print('Updating Spreadsheet...')
    info = parse_page(url)

    input_info('20' + date_finished[-2:], info)
    input_info('Overall', info)

    wb.save('Booktracker.xlsx')

    print('Booktracker has completed and will now close.')
