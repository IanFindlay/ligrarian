#!/usr/bin/env python3

"""Do a test run of the Ligrarian that resets everything afterwards."""

from datetime import datetime as dt
from datetime import timedelta
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait

import ligrarian


def review_function_cluster(driver, book_info, reread, shelves, url):
    """Run through the Goodreads review functions."""
    ligrarian.goodreads_date_input(driver, book_info['date'], reread)
    ligrarian.goodreads_add_review(driver, book_info['review'])
    driver.find_element_by_name('next').click()
    driver.get(url)
    ligrarian.goodreads_rate_book(driver, book_info['rating'])
    if not reread:
        ligrarian.goodreads_shelve(driver, shelves)


def run_test():
    """Perform the Ligrarian test run."""
    email = ligrarian.get_setting('User', 'Email')
    password = ligrarian.get_setting('User', 'Password')
    yesterday = dt.strftime(dt.now() - timedelta(1), '%d/%m/%y')
    today = dt.strftime(dt.now(), '%d/%m/%y')
    book_info = {
        'terms': 'Cannery Row', 'date': yesterday,
        'format': 'kindle', 'rating': '4', 'review': 'Test Review',
    }

    driver = webdriver.Firefox()
    driver.implicitly_wait(10)

    ligrarian.goodreads_login(driver, email, password)
    ligrarian.goodreads_find(driver, book_info['terms'])
    url = ligrarian.goodreads_filter(driver, book_info['format'])
    shelves = ligrarian.goodreads_get_shelves(driver, book_info['rating'])

    # Check book is correct format
    info_rows = driver.find_elements_by_class_name('row')
    row_text = [row.text for row in info_rows]
    assert 'Kindle' in ''.join(row_text), "Book is in incorrect format."

    # Test review of a book that hasn't been read by the account before
    review_function_cluster(driver, book_info, False, shelves, url)

    undo_elem = driver.find_element_by_class_name('wtrStatusRead.wtrUnshelve')
    assert undo_elem, "Book wasn't marked as read."
    print("Book successfully marked as read.")

    # Need a delay to avoid a Goodreads error popup
    time.sleep(3)
    # Change date and run testing for rereading a book
    book_info['date'] = today

    review_function_cluster(driver, book_info, True, shelves, url)

    # Reset Goodreads account
    driver.get(url)

    # Reset Goodreads account - Have to redo the find for undo due to refreshes
    undo_elem = driver.find_element_by_class_name('wtrStatusRead.wtrUnshelve')
    undo_elem.click()
    alert_obj = driver.switch_to.alert
    alert_obj.accept()

    # Check book no longer marked as read before closing window
    unread_check = driver.find_element_by_class_name('wtrToRead')
    assert unread_check, "Problem resetting account - Test run specific error."
    print('Goodreads account reset correctly.')
    driver.close()

    # Spreadsheet entry testing
    path = ligrarian.get_setting('Settings', 'Path')
    workbook = openpyxl.load_workbook(path)
    print('Testing spreadsheet updating.')
    info = ligrarian.parse_page(url)
    info['category'], info['genre'] = ligrarian.category_and_genre(shelves)
    print(info)

    year_sheet = '20' + book_info['date'][-2:]

    # Get first blank row before attempting to write
    pre_year = ligrarian.first_blank_row(workbook[year_sheet])
    pre_overall = ligrarian.first_blank_row(workbook['Overall'])
    workbook.close()

    # Try to write info to both sheets
    ligrarian.input_info(year_sheet, info, book_info['date'])

    # Find blank rows now that the data should have been entered
    workbook = openpyxl.load_workbook(path)
    post_year = ligrarian.first_blank_row(workbook[year_sheet])
    post_overall = ligrarian.first_blank_row(workbook['Overall'])

    # Confirm data was entered
    assert pre_year < post_year, "Data not written to year sheet."
    assert pre_overall < post_overall, "Data not written to overall sheet."
    print("Data was written to sheet successfully.")

    # Delete newly entered data
    for sheet in [year_sheet, 'Overall']:
        sheet = workbook[sheet]

        input_row = ligrarian.first_blank_row(sheet) - 1

        sheet.cell(row=input_row, column=1).value = ''
        sheet.cell(row=input_row, column=2).value = ''
        sheet.cell(row=input_row, column=3).value = ''
        sheet.cell(row=input_row, column=4).value = ''
        sheet.cell(row=input_row, column=5).value = ''
        sheet.cell(row=input_row, column=6).value = ''

    workbook.save(path)

    print("Spreadsheet reset.")
    print("Test run complete.")


if __name__ == '__main__':
    run_test()
