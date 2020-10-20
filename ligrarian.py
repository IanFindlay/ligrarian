#!/usr/bin/env python3

"""Automatically update Goodreads and local Spreadsheet with book read info.

Args:
    Three operational modes (g)ui, (s)earch or (u)rl

    gui arguments:
        None

    search arguments:
        Title of Book: Enclosed in double quotation marks
        Author of Book: Enclosed in double quotation marks
        Format: (p)aperback, (h)ardcover, (k)indle or (e)book
        Read Date: (t)oday, (y)esterday or a date formatted DD/MM/YY
        Rating: Number between 1 and 5
        Review (Optional): Enclosed in double quotation marks

    url arguments:
        URL: Goodreads URL for the book
        Read Date: (t)oday, (y)esterday or a date formatted DD/MM/YY
        Rating: Number between 1 and 5
        Review (Optional): Enclosed in double quotation marks
"""

import argparse
import configparser
from datetime import datetime as dt
from datetime import timedelta
import sys
import tkinter as tk
from tkinter import messagebox

import bs4
import openpyxl
import requests
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait


class Gui:
    """Acts as the base of the GUI and contains the assoicated methods."""

    def __init__(self, master, settings):
        """Gui class constructor to initialise Gui object.

        Args:
            master (obj): tkinter TK object - base object for the GUI.

        """
        self.master = master
        self.master.title("Ligrarian")
        self.master.geometry('665x560')
        self.settings = settings
        self.info = {}

        # Labels
        login_label = tk.Label(self.master, text="Login")
        login_label.grid(row=2, column=1, sticky='W',
                         pady=(30, 5), padx=10)

        self.main_label = tk.Label(self.master, text='Search')
        self.main_label.grid(row=3, column=1, sticky='W', padx=10)

        date_label = tk.Label(self.master, text="Date")
        date_label.grid(row=5, column=1, sticky='W', padx=10)

        self.format_label = tk.Label(self.master, text="Format")
        self.format_label.grid(row=6, column=1, sticky='W', padx=10)

        rating_label = tk.Label(self.master, text="Rating")
        rating_label.grid(row=7, column=1, sticky='W', padx=10)

        review_label = tk.Label(
            self.master, text="Review\n (optional)", padx=10
        )
        review_label.grid(row=8, column=1, sticky='W')

        # Widgets
        self.email = tk.Entry(self.master, width=20)
        email = self.settings['email']
        if email:
            self.email.insert(0, email)
        else:
            self.email.insert(0, 'Email')
        self.email.grid(row=2, column=2, columnspan=3,
                        sticky='W', pady=(30, 5))

        self.password = tk.Entry(self.master, width=20)
        password = self.settings['password']
        if password:
            self.password.insert(0, '********')
        else:
            self.password.insert(0, 'Password')
        self.password.grid(row=2, column=4, columnspan=3, sticky='W',
                           pady=(30, 5))

        self.save_choice = tk.IntVar()
        save_box = tk.Checkbutton(
            self.master, text='Save Password', variable=self.save_choice,
            onvalue=True, offvalue=False
        )
        save_box.grid(row=2, column=7, sticky='W', pady=(30, 5))
        if password:
            save_box.select()

        self.main = tk.Entry(self.master, width=45)
        self.main.grid(row=3, column=2, columnspan=6, sticky='W', pady=10)

        self.mode = tk.IntVar()
        mode_button = tk.Checkbutton(
            self.master, text='URL Mode', variable=self.mode,
            onvalue=True, offvalue=False,
            command=self.mode_switch
        )
        mode_button.grid(row=3, column=7, sticky='W')

        formats = ("Paperback", "Hardback", "Kindle", "Ebook",)
        self.format = tk.StringVar()
        self.format.set(self.settings["format"])
        self.format_menu = tk.OptionMenu(self.master, self.format, *formats)
        self.format_menu.grid(row=6, column=2, columnspan=3,
                              sticky='W', pady=5)

        self.date = tk.Entry(self.master, width=8)
        self.date.insert(0, 'DD/MM/YY')
        self.date.grid(row=5, column=2, sticky='W', pady=10, ipady=3, ipadx=5)

        today_button = tk.Button(self.master, text="Today",
                                 command=self.set_date)

        today_button.grid(row=5, column=3, sticky='W', pady=10,)

        yesterday_button = tk.Button(self.master, text="Yesterday",
                                     command=lambda: self.set_date(True))
        yesterday_button.grid(row=5, column=4, sticky='W', pady=10)

        stars = ("1", "2", "3", "4", "5")
        self.rating = tk.StringVar()
        self.rating.set(self.settings["rating"])
        rating_menu = tk.OptionMenu(self.master, self.rating, *stars)
        rating_menu.grid(row=7, column=2, sticky='W', pady=5)

        self.review = tk.Text(self.master, height=15,
                              width=75, wrap=tk.WORD)
        self.review.grid(row=8, column=2, columnspan=7, sticky='W', pady=5)

        submit_button = tk.Button(self.master, text="Mark as Read",
                                  command=self.parse_input)
        submit_button.grid(row=12, column=7, columnspan=2, sticky='E', pady=15)

    def mode_switch(self):
        """Modify displayed widgets and edit label text depending on mode."""
        if self.mode.get():
            self.format_label.grid_remove()
            self.format_menu.grid_remove()
            self.main_label.configure(text='URL')
        else:
            self.format_label.grid()
            self.format_menu.grid()
            self.main_label.configure(text='Search')

    def set_date(self, yesterday=False):
        """Set date widget to a new value.

        Args:
            yesterday (bool): Whether to set date to yesterday's date or not.

        """
        self.date.delete(0, 8)
        self.date.insert(0, get_date_str(yesterday))

    def parse_input(self):
        """Create input dictionary and test required info has been given."""
        self.mode = self.mode.get()
        self.settings['email'] = self.email.get()
        password = self.password.get()
        if save_choice.get():
            self.settings['password'] = password

        if password == '********':
            password = self.settings['password']

        self.info = {
            'main': self.main.get(),
            'date': self.date.get(),
            'format': self.format.get(),
            'rating': self.rating.get(),
            'review': self.review.get('1.0', 'end-1c'),
        }

        try:
            assert self.settings['email'] != 'Email'
            assert self.settings['password'] != 'Password'
            assert self.info['main']
            if not self.mode:
                assert self.info['format']
            self.master.destroy()

        except AssertionError:
            messagebox.showwarning(message="Complete all non-optional "
                                   "fields before marking as read."
            )


def retrieve_settings():
    """Retrieve settings from .ini and return values as dictionary.

    Returns:
        settings (dict): Dictionary with option: value format.
    """
    config = configparser.ConfigParser()
    config.read('settings.ini')
    settings = {}
    for section in config.sections():
        for key, value in config.items(section):
            if key in ['prompt', 'headless']:
                value = bool(value)
            settings[key] = value

    return settings


def create_gui(settings_dict):
    """Create GUI instance and return it."""
    root = tk.Tk()
    root.protocol("WM_DELETE_WINDOW", exit)
    gui = Gui(root, settings_dict)
    root.mainloop()

    return gui


def gui_mode_details_edits(gui):
    """Assess gui mode and edit details dictionary accordingly.

    Args:
        gui (obj): Instance of GUI class.
    """
    if gui.mode:
        gui.info['url'] = gui.info['main']
    else:
        gui.info['search'] = gui.info['main']
    del gui.info['main']

    return gui.info


def get_date_str(yesterday=False):
    """Return a string of today's or yesterday's date.

    Args:
        yesterday (bool): Whether to get yesterday's date or not.

    Returns:
        Strftime datetime of today's or yesterday's date formatted 'DD/MM/YY'.

    """
    today_datetime = dt.now()
    if yesterday:
        return dt.strftime(today_datetime - timedelta(1), '%d/%m/%Y')
    return dt.strftime(today_datetime, '%d/%m/%Y')


def parse_arguments():
    """Set up parsers/subparsers and parse command line arguments.

    Returns:
        Dictionary of parsed arguments.

    """
    parser = argparse.ArgumentParser(description="Goodreads updater")
    subparsers = parser.add_subparsers(help="Choose (u)rl, (s)earch or (g)ui")

    url_parser = subparsers.add_parser("url", aliases=['u'])
    url_parser.add_argument('url', metavar="url",
                            help="Book's Goodreads URL within quotes")
    url_parser.add_argument('date', help=("(t)oday, (y)esterday or "
                                          "date formatted DD/MM/YY"))
    url_parser.add_argument('rating', type=int, metavar='rating',
                            choices=[1, 2, 3, 4, 5],
                            help="A number 1 through 5")
    url_parser.add_argument('review', nargs='?', metavar="'review'",
                            help="Review enclosed in quotes")

    search_parser = subparsers.add_parser('search', aliases=['s'])
    search_parser.add_argument('search', metavar="'search terms'",
                               help="Search terms to use e.g. Book title "
                                    "and Author")
    search_parser.add_argument('format', metavar='format',
                               choices=['e', 'h', 'k', 'p'],
                               help="(p)aperback, (h)ardcover, "
                                    "(k)indle, (e)book")
    search_parser.add_argument('date', help=("(t)oday, (y)esterday or "
                                             "date formatted DD/MM/YY"))
    search_parser.add_argument('rating', type=int, metavar='rating',
                               choices=[1, 2, 3, 4, 5],
                               help="A number 1 through 5")
    search_parser.add_argument('review', nargs='?', metavar="'review'",
                               help="Review enclosed in quotes")

    gui = subparsers.add_parser("gui", aliases=['g'])
    gui.add_argument('gui', action='store_true',
                     help="Invoke GUI (Defaults to True)")

    args = parser.parse_args()
    return vars(args)


def create_driver(run_headless):
    """Create the appropriate driver for the session.

    Args:
        run_headless (bool): Run in headless mode or not
    """
    if run_headless:
        print(('Opening a headless computer controlled browser and updating '
               'Goodreads'))
        options = Options()
        options.headless = True
        return webdriver.Firefox(options=options)
    print('Opening a computer controlled browser and updating Goodreads')
    return webdriver.Firefox()


def check_and_prompt_for_email_password(settings_dict):
    """Assess if email and password are missing and if so prompt for them.

    Args:
        settings (dict): Dictionary of user settings
    """
    if not settings_dict['email']:
        settings_dict['email'] = input('Email: ')

    if not settings_dict['password']:
        password = input('Password: ')
        if settings_dict['prompt']:
            save = input("Save Password?(y/n): ")
            if save.lower() == 'y':
                settings_dict['password'] = password
            elif save.lower() == 'n':
                disable = input("Disable save Password prompt?(y/n): ")
                if disable.lower() == 'y':
                    settings_dict['prompt'] = False
                else:
                    settings_dict['prompt'] = True


def write_initial_config():
    """Writes the inital, default settings.ini file."""
    config = configparser.ConfigParser()
    config['user'] = {'email': '',
                      'password': ''}
    config['settings'] = {'prompt': 'False',
                          'path': './Ligrarian.xlsx',
                          'headless': 'False'}
    config['defaults'] = {'format': 'Paperback',
                          'rating': '3'}
    with open('settings.ini', 'w') as configfile:
        config.write(configfile)


def write_config(email, password, prompt):
    """Write configuration file.

    Args:
        email (str): Email address to write to the config file.
        password (str): Password to write to the config file.
        prompt (str): Whether to disable the save prompt ('yes' or 'no').

    """
    config = configparser.ConfigParser()
    config.read('settings.ini')
    config.set('user', 'email', email)
    config.set('user', 'password', password)
    config.set('settings', 'prompt', prompt)

    with open('settings.ini', 'w') as configfile:
        config.write(configfile)


def goodreads_login(driver, email, password):
    """Login to Goodreads account from the homepage.

    Args:
        driver: Selenium webdriver to act upon.
        email (str): Email address to be entered.
        password (str): Password to be entered.

    """
    driver.get('https://goodreads.com')

    driver.find_element_by_name('user[email]').send_keys(email)
    pass_elem = driver.find_element_by_name('user[password]')
    pass_elem.send_keys(password, Keys.ENTER)

    try:
        driver.find_element_by_class_name('siteHeader__personal')
    except NoSuchElementException:
        print('Failed to login - Email and/or Password probably incorrect.')
        driver.close()
        sys.exit()


def goodreads_find(driver, terms):
    """Find the book on Goodreads and navigate to all editions page.

    Args:
        driver: Selenium webdriver to act upon.
        terms (str): Terms to be used in the Goodreads search.

    Raises:
        NoSuchElementException: Search terms yields no results.

    """
    search_elem = driver.find_element_by_class_name('searchBox__input')
    search_elem.send_keys(terms, Keys.ENTER)

    try:
        driver.find_element_by_partial_link_text('edition').click()
    except NoSuchElementException:
        print("Failed to find book using those search terms.")
        driver.close()
        sys.exit()


def goodreads_filter(driver, book_format):
    """Filter editions with book_format and select top book.

    Args:
        driver: Selenium webdriver to act upon.
        book_format (str): The format of the book.

    Returns:
        Current URL the driver argument is now visiting.

    """
    pre_filter_url = driver.current_url

    # Filter by format
    filter_elem = driver.find_element_by_name('filter_by_format')
    filter_elem.click()
    filter_elem.send_keys(book_format, Keys.ENTER)

    # Make sure filtered page is loaded before clicking top book
    WebDriverWait(driver, 10).until(
        EC.url_changes((pre_filter_url))
    )

    # Select top book
    driver.find_element_by_class_name('bookTitle').click()

    return driver.current_url


def goodreads_get_shelves(driver, rating):
    """Find and return list of 'Top Shelves' on Goodreads book page.

    Args:
        driver: Selenium webdriver to act upon.
        rating (str): String representation of a number 1-5.

    Returns:
        list of strings of the 'shelve' categories on the current driver page.

    """
    shelves_elems = driver.find_elements_by_class_name('actionLinkLite.'
                                                       'bookPageGenreLink')
    shelves = []
    for shelf in shelves_elems:
        if ' users' not in shelf.text and shelf.text not in shelves:
            shelves.append(shelf.text)

    if rating == '5':
        shelves.append('5-star-books')

    return shelves


def goodreads_date_input(driver, date_done, reread):
    """Select completion date on review page, add new selectors for rereads.

    Args:
        driver: Selenium webdriver to act upon.
        date_done (str): Date formatted DD/MM/YY.
        Reread (str or None): String of a number 1-5, essentially acting as
                              the boolean True, or None acting as False.

    """
    book_code = driver.current_url.split('/')[-1]
    driver.get("https://www.goodreads.com/review/edit/{}".format(book_code))

    # If it's a reread need to create new session selectors
    if reread:
        driver.find_element_by_id('readingSessionAddLink').click()
        # More details loaded for Explicit Wait
        driver.find_element_by_class_name('smallLink.closed').click()
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.ID, "review_recommendation"))
        )

    # Find reading session codes from all ids then use last one for date entry
    reread_codes = []
    for id_elem in driver.find_elements_by_xpath('//*[@id]'):
        id_attribute = id_elem.get_attribute('id')
        if 'readingSessionEntry' in id_attribute:
            reread_codes.append(id_attribute)
    new_read_code = reread_codes[-1][19:]

    # Date Formatting and Selection
    year = date_done[6:]
    month = date_done[3:5].lstrip('0')
    day = date_done[:2].lstrip('0')

    year_name = 'readingSessionDatePicker{}[end][year]'.format(new_read_code)
    month_name = 'readingSessionDatePicker{}[end][month]'.format(new_read_code)
    day_name = 'readingSessionDatePicker{}[end][day]'.format(new_read_code)

    Select(driver.find_element_by_name(year_name)
           ).select_by_visible_text(year)

    Select(driver.find_element_by_name(month_name)
           ).select_by_value(month)

    Select(driver.find_element_by_name(day_name)
           ).select_by_visible_text(day)


def goodreads_add_review(driver, review):
    """Write review in review box (if given).

    Args:
        driver: Selenium webdriver to act upon.
        review (str): Review of the book.

    """
    review_elem = driver.find_element_by_name('review[review]')
    review_elem.clear()
    review_elem.click()
    review_elem.send_keys(review)


def goodreads_rate_book(driver, rating):
    """Give the book the given rating out of 5.

    Args:
        driver: Selenium webdriver to act upon.
        rating (str): A number 1-5.

    """
    # Give star rating
    stars_elem = driver.find_elements_by_class_name('star.off')
    for stars in stars_elem:
        if stars.text.strip() == '{} of 5 stars'.format(rating):
            stars.click()
            break


def goodreads_shelve(driver, shelves):
    """Add the Goodreads book to relevant user shelves.

    Use the list of 'Top Shelves' and a users own active shelves to add the
    book on the driver's current page to each of the relevant shelves.

    Args:
        driver: Selenium webdriver to act upon.
        shelves (list): List of strings representing Goodreads shelves.

    """
    # Wait until review box is invisible
    WebDriverWait(driver, 10).until(
        EC.invisibility_of_element_located((By.ID, "box"))
    )

    # Select shelves
    menu_elem = driver.find_element_by_class_name('wtrShelfButton')
    menu_elem.click()
    shelf_elem = driver.find_element_by_class_name('wtrShelfSearchField')

    for shelf in shelves:
        shelf_elem.send_keys(shelf, Keys.ENTER)
        shelf_elem.send_keys(Keys.SHIFT, Keys.HOME, Keys.DELETE)

    # Close dropdown and wait until it disappears
    menu_elem.click()
    WebDriverWait(driver, 10).until(
        EC.invisibility_of_element_located((By.CLASS_NAME, "wtrShelfList"))
    )


def parse_page(url):
    """Parse Goodreads page for title, author and number of pages.

    Args:
        url (str): Goodreads Book URL.

    Returns:
        Dictionary of parsed Title, Author and Number of Pages.

    """
    info = {}
    res = requests.get(url)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    title_elem = soup.select('#bookTitle')
    rough_title = title_elem[0].getText().strip().split('\n')
    if len(rough_title) == 1:
        info['title'] = rough_title[0].strip()
    else:
        info['title'] = rough_title[0].strip() + ' ' + rough_title[2].strip()

    info['author'] = soup.select('.authorName')[0].getText().strip()

    pages_elem = soup.findAll('span', attrs={'itemprop': 'numberOfPages'})
    info['pages'] = int(pages_elem[0].getText().strip(' pages'))

    return info


def category_and_genre(shelves):
    """Use shelves list to deterime genre and categorise as Fiction/Nonfiction.

    Args:
        shelves (list): List of strings of Goodreads shelf categories.

    Returns:
        Tuple of category (Fiction/Nonfiction) and the books genre.

    """
    if 'Nonfiction' in shelves:
        category = 'Nonfiction'
    else:
        category = 'Fiction'

    for shelf in shelves:
        if shelf != category:
            genre = shelf
            break

    return (category, genre)


def input_info(path, year_sheet, info, date):
    """Write the book information to the first blank row on the given sheet.

    Args:
        year_sheet (str): The year the book was read formatted YYYY.
        info (dict): Information about the book.
        date (str): Date to input in the 'Read date' column.

    """
    workbook = openpyxl.load_workbook(path)

    existing_sheets = workbook.sheetnames
    if year_sheet not in existing_sheets:
        create_sheet(workbook, existing_sheets[-1], year_sheet)

    for sheet in [year_sheet, 'Overall']:
        sheet = workbook[sheet]

        input_row = first_blank_row(sheet)

        sheet.cell(row=input_row, column=1).value = info['title']
        sheet.cell(row=input_row, column=2).value = info['author']
        sheet.cell(row=input_row, column=3).value = info['pages']
        sheet.cell(row=input_row, column=4).value = info['category']
        sheet.cell(row=input_row, column=5).value = info['genre']
        sheet.cell(row=input_row, column=6).value = date

    workbook.save(path)


def create_sheet(workbook, sheet_to_copy, new_sheet_name):
    """Create a new sheet by copying and modifying a different one.

    Args:
        workbook (obj): openpyxl workbook object.
        sheet_to_copy (str): Name of the sheet to copy.
        new_sheet_name (str): Name (year formatted YYYY) to name the new sheet.

    """
    sheet = workbook.copy_worksheet(workbook[sheet_to_copy])
    sheet.title = new_sheet_name
    last_row = first_blank_row(sheet)
    while last_row > 1:
        for col in range(1, 7):
            sheet.cell(row=last_row, column=col).value = None
        last_row -= 1
    day_tracker = '=(TODAY()-DATE({},1,1))/7'.format(new_sheet_name)
    sheet.cell(row=5, column=9).value = day_tracker


def first_blank_row(sheet):
    """Return the number of the first blank row of the given sheet.

    Args:
        sheet (obj): openpyxl sheet object to find first blank row of.

    """
    input_row = 0
    while True:
        input_row += 1
        if not sheet.cell(row=input_row, column=1).value:
            break
    return input_row


def main():
    """Coordinate updating of Goodreads account and writing to spreadsheet."""
    args = parse_arguments()
    try:
        open("settings.ini")
    except FileNotFoundError:
        write_initial_config()

    settings = retrieve_settings()

    if 'gui' in args:
        gui_instance = create_gui()
        details = gui_mode_details_edits(gui_instance)

    else:
        details = args
        check_and_prompt_for_email_password(settings)
        # Process date if given as (t)oday or (y)esterday into proper format
        if details['date'].lower() == 't':
            details['date'] = get_date_str()
        elif details['date'].lower() == 'y':
            details['date'] = get_date_str(True)

    driver = create_driver(settings['headless'])

    driver.implicitly_wait(10)
    goodreads_login(driver, settings['email'], settings['password'])
    if 'url' in details:
        url = details['url']
        driver.get(url)
    else:
        goodreads_find(driver, details['search'])
        url = goodreads_filter(driver, details['format'])

    shelves = goodreads_get_shelves(driver, details['rating'])

    # Use rating element to see if book has been read before
    rating_elem = driver.find_element_by_class_name('stars')
    reread = rating_elem.get_attribute('data-rating') != '0'

    goodreads_date_input(driver, details['date'], reread)

    if details['review']:
        goodreads_add_review(driver, details['review'])

    driver.find_element_by_name('next').click()
    driver.get(url)
    goodreads_rate_book(driver, details['rating'])

    if not reread:
        goodreads_shelve(driver, shelves)

    driver.close()
    print('Goodreads account updated.')

    print('Updating Spreadsheet...')
    info = parse_page(url)
    info['category'], info['genre'] = category_and_genre(shelves)
    year_sheet = '20' + details['date'][-2:]
    input_info(settings['path'], year_sheet, info, details['date'])

    print(('Ligrarian has completed and will now close. The following '
           'information has been written to the spreadsheet:'))
    print(info['title'], info['author'], info['pages'],
          info['category'], info['genre'], details['date'], sep='\n')

    write_config(settings['email'], settings['password'], settings['prompt'])


if __name__ == '__main__':
    main()
